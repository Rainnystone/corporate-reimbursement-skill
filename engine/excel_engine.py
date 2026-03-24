import openpyxl
from copy import copy
from .utils import setup_logger

logger = setup_logger(__name__)

def safe_write(worksheet, coordinate, value):
    """
    Safely writes a value to a cell, protecting existing formulas.
    If the cell is merged, it attempts to write to the top-left cell of the range.
    """
    cell = worksheet[coordinate]
    
    # Check if the cell is part of a merged range
    for merged_range in worksheet.merged_cells.ranges:
        if coordinate in merged_range:
            # If it's the top-left cell, we can write directly
            if coordinate == merged_range.coord.split(':')[0]:
                break
            else:
                # If it's not the top-left, we write to the top-left instead
                top_left = merged_range.coord.split(':')[0]
                logger.info(f"Cell {coordinate} is merged. Writing to top-left cell {top_left} instead.")
                cell = worksheet[top_left]
                break
    
    if cell.data_type == 'f':
        logger.warning(f"Attempted to overwrite formula in {coordinate} of {worksheet.title}. Skipped.")
        return False
    
    cell.value = value
    return True

class ReimbursementExcel:
    def __init__(self, template_path):
        self.template_path = template_path
        self.workbook = openpyxl.load_workbook(template_path, data_only=False)
        
    def insert_row_with_formats(self, sheet_name, row_idx):
        """
        Inserts a row at row_idx and copies formats/styles from the row above it.
        This is crucial when a user has more entries than pre-allocated rows.
        """
        ws = self.workbook[sheet_name]
        ws.insert_rows(row_idx)
        logger.info(f"Inserted new row at index {row_idx} in sheet '{sheet_name}'.")
        
        # Copy style from row above
        if row_idx > 1:
            for col_idx in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=row_idx - 1, column=col_idx)
                target_cell = ws.cell(row=row_idx, column=col_idx)
                if source_cell.has_style:
                    target_cell.font = openpyxl.styles.copy(source_cell.font) if hasattr(openpyxl.styles, 'copy') else copy(source_cell.font)
                    target_cell.border = openpyxl.styles.copy(source_cell.border) if hasattr(openpyxl.styles, 'copy') else copy(source_cell.border)
                    target_cell.fill = openpyxl.styles.copy(source_cell.fill) if hasattr(openpyxl.styles, 'copy') else copy(source_cell.fill)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = openpyxl.styles.copy(source_cell.protection) if hasattr(openpyxl.styles, 'copy') else copy(source_cell.protection)
                    target_cell.alignment = openpyxl.styles.copy(source_cell.alignment) if hasattr(openpyxl.styles, 'copy') else copy(source_cell.alignment)
        
        return True

    def save(self, output_path):
        self.workbook.save(output_path)
        logger.info(f"Saved generated excel to {output_path}")
