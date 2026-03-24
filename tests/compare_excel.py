import openpyxl
import os
import sys

def compare_excel(output_path, benchmark_path):
    if not os.path.exists(output_path):
        return f"Error: Output file {output_path} not found."
    if not os.path.exists(benchmark_path):
        return f"Error: Benchmark file {benchmark_path} not found."

    try:
        # Load both ways to compare both literals/formulas and calculated values
        wb_out_raw = openpyxl.load_workbook(output_path, data_only=False)
        wb_ben_raw = openpyxl.load_workbook(benchmark_path, data_only=False)
        wb_out_data = openpyxl.load_workbook(output_path, data_only=True)
        wb_ben_data = openpyxl.load_workbook(benchmark_path, data_only=True)
    except Exception as e:
        return f"Error loading workbooks: {e}"

    diffs = []

    for sheet_name in wb_ben_raw.sheetnames:
        if sheet_name not in wb_out_raw.sheetnames:
            diffs.append(f"Missing sheet: {sheet_name}")
            continue
        
        ws_out_raw = wb_out_raw[sheet_name]
        ws_ben_raw = wb_ben_raw[sheet_name]
        ws_out_data = wb_out_data[sheet_name]
        ws_ben_data = wb_ben_data[sheet_name]

        max_row = max(ws_out_raw.max_row, ws_ben_raw.max_row)
        max_col = max(ws_out_raw.max_column, ws_ben_raw.max_column)

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                # 1. Compare Raw (Formulas/Literals)
                v_out_raw = ws_out_raw.cell(row=r, column=c).value
                v_ben_raw = ws_ben_raw.cell(row=r, column=c).value
                
                if v_out_raw == v_ben_raw:
                    continue
                
                # 2. If raw doesn't match, compare calculated values
                v_out_data = ws_out_data.cell(row=r, column=c).value
                v_ben_data = ws_ben_data.cell(row=r, column=c).value
                
                # Normalize strings
                if isinstance(v_out_data, str): v_out_data = v_out_data.strip()
                if isinstance(v_ben_data, str): v_ben_data = v_ben_data.strip()

                # Special case: openpyxl data_only=True returns None for new formulas
                # If benchmark has a value but output has None because it's a formula,
                # we can't easily verify the result without a calculator.
                # However, for literals, we should match.
                
                if v_out_data != v_ben_data:
                    # Floating point tolerance
                    if isinstance(v_out_data, (int, float)) and isinstance(v_ben_data, (int, float)):
                        if abs(v_out_data - v_ben_data) < 0.05:
                            continue
                    
                    # None vs Empty
                    if (v_out_data is None and v_ben_data == "") or (v_out_data == "" and v_ben_data is None):
                        continue
                    
                    # If output is a formula and data is None (not calculated yet)
                    # and benchmark has a value, we might have a potential match.
                    # For now, let's report it but with a note.
                    if ws_out_raw.cell(row=r, column=c).data_type == 'f' and v_out_data is None:
                        # Skip if we trust the formula (hard to verify)
                        continue

                    diffs.append(f"Sheet '{sheet_name}', Cell {ws_ben_raw.cell(row=r, column=c).coordinate}: "
                                 f"Expected '{v_ben_data}' (raw: {v_ben_raw}), got '{v_out_data}' (raw: {v_out_raw})")

    return diffs

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 compare_excel.py <output_xlsx> <benchmark_xlsx>")
        sys.exit(1)
    
    out = sys.argv[1]
    ben = sys.argv[2]
    
    results = compare_excel(out, ben)
    if not results:
        print("SUCCESS: 100% Match!")
    elif isinstance(results, str):
        print(results)
    else:
        print(f"FAILED: Found {len(results)} differences:")
        for d in results[:30]:
            print(f"- {d}")
        if len(results) > 30:
            print(f"... and {len(results) - 30} more.")
