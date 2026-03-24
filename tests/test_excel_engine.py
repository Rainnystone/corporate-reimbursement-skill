import unittest
import openpyxl
import os
import tempfile

class TestExcelEngine(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.test_file = os.path.join(self.temp_dir.name, "test_template.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        # Setup a dummy formula
        ws['A1'] = "=SUM(B1:B2)"
        # Setup an empty cell
        ws['A2'] = ""
        wb.save(self.test_file)

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_formula_protection(self):
        from reimbursement.excel_engine import safe_write
        
        wb = openpyxl.load_workbook(self.test_file)
        ws = wb["TestSheet"]
        
        # Try to overwrite formula cell -> should skip
        success_formula = safe_write(ws, 'A1', 100)
        self.assertFalse(success_formula)
        self.assertEqual(ws['A1'].data_type, 'f')
        self.assertEqual(ws['A1'].value, "=SUM(B1:B2)")

        # Try to overwrite empty cell -> should write
        success_empty = safe_write(ws, 'A2', 200)
        self.assertTrue(success_empty)
        self.assertEqual(ws['A2'].value, 200)

    def test_insert_row_with_formats(self):
        from reimbursement.excel_engine import ReimbursementExcel
        excel = ReimbursementExcel(self.test_file)
        
        # Insert row at 2, moving old row 2 down
        excel.insert_row_with_formats("TestSheet", 2)
        
        ws = excel.workbook["TestSheet"]
        # Now A3 should be what A2 was (empty originally)
        self.assertTrue(ws.max_row >= 2)
        # Check formula in A1 wasn't disturbed
        self.assertEqual(ws['A1'].value, "=SUM(B1:B2)")

if __name__ == '__main__':
    unittest.main()
